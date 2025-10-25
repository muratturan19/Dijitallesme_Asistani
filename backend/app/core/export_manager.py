# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportManager:
    """Handles Excel export of extracted data"""

    def __init__(self, output_dir: Path):
        """
        Initialize export manager

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(
        self,
        template_fields: List[Dict[str, Any]],
        extracted_data_list: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        Export extracted data to Excel file

        Args:
            template_fields: Template field definitions
            extracted_data_list: List of extracted data records
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Çıkarılan Veriler"

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.xlsx"

            # Add header row
            headers = [field['field_name'] for field in template_fields]
            ws.append(headers)

            # Style header row
            self._style_header_row(ws, len(headers))

            # Add data rows
            for data_record in extracted_data_list:
                row_data = []
                field_values = data_record.get('field_values', {})

                for field in template_fields:
                    field_name = field['field_name']
                    value = field_values.get(field_name)

                    # Format value based on data type
                    formatted_value = self._format_value(
                        value,
                        field.get('data_type', 'text')
                    )

                    row_data.append(formatted_value)

                ws.append(row_data)

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)

            # Add metadata sheet
            self._add_metadata_sheet(wb, template_fields, extracted_data_list)

            # Save file
            output_path = self.output_dir / filename
            wb.save(str(output_path))

            logger.info(f"Excel dosyası oluşturuldu: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Excel export hatası: {str(e)}")
            raise

    def _style_header_row(self, ws, num_columns: int):
        """
        Apply styling to header row

        Args:
            ws: Worksheet
            num_columns: Number of columns
        """
        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

    def _auto_adjust_columns(self, ws):
        """
        Auto-adjust column widths based on content

        Args:
            ws: Worksheet
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_value(self, value: Any, data_type: str) -> Any:
        """
        Format value based on data type

        Args:
            value: Raw value
            data_type: Field data type

        Returns:
            Formatted value
        """
        if value is None:
            return ""

        try:
            if data_type == 'number':
                # Try to convert to number
                if isinstance(value, str):
                    # Handle Turkish number format (1.234,56)
                    value = value.replace('.', '').replace(',', '.')
                return float(value)

            elif data_type == 'date':
                # Keep as string for now (could parse to datetime)
                return str(value)

            else:  # text
                return str(value)

        except:
            return str(value)

    def _add_metadata_sheet(
        self,
        wb,
        template_fields: List[Dict[str, Any]],
        extracted_data_list: List[Dict[str, Any]]
    ):
        """
        Add metadata sheet with export information

        Args:
            wb: Workbook
            template_fields: Template field definitions
            extracted_data_list: Extracted data records
        """
        try:
            # Create metadata sheet
            ws_meta = wb.create_sheet("Metadata")

            # Add export info
            ws_meta.append(["Dışa Aktarma Bilgileri"])
            ws_meta.append([])
            ws_meta.append(["Tarih:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
            ws_meta.append(["Kayıt Sayısı:", len(extracted_data_list)])
            ws_meta.append(["Alan Sayısı:", len(template_fields)])
            ws_meta.append([])

            # Add field information
            ws_meta.append(["Alan Bilgileri"])
            ws_meta.append(["Alan Adı", "Veri Tipi", "Zorunlu"])

            for field in template_fields:
                ws_meta.append([
                    field['field_name'],
                    field.get('data_type', 'text'),
                    "Evet" if field.get('required', False) else "Hayır"
                ])

            # Style metadata sheet
            ws_meta['A1'].font = Font(bold=True, size=14)

        except Exception as e:
            logger.warning(f"Metadata sheet oluşturma hatası: {str(e)}")

    def export_validation_report(
        self,
        extracted_data_list: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        Export validation report with confidence scores

        Args:
            extracted_data_list: List of extracted data with confidence
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Doğrulama Raporu"

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"validation_report_{timestamp}.xlsx"

            # Add headers
            headers = ["Belge", "Alan", "Değer", "Güven Skoru", "Durum"]
            ws.append(headers)
            self._style_header_row(ws, len(headers))

            # Add data
            for idx, data_record in enumerate(extracted_data_list, 1):
                doc_name = data_record.get('document_name', f'Belge {idx}')
                field_values = data_record.get('field_values', {})
                confidence_scores = data_record.get('confidence_scores', {})

                for field_name, value in field_values.items():
                    confidence = confidence_scores.get(field_name, 0.0)

                    # Determine status
                    if confidence >= 0.8:
                        status = "Yüksek"
                        fill_color = "C6EFCE"  # Green
                    elif confidence >= 0.5:
                        status = "Orta"
                        fill_color = "FFEB9C"  # Yellow
                    else:
                        status = "Düşük"
                        fill_color = "FFC7CE"  # Red

                    ws.append([doc_name, field_name, value, f"{confidence:.2%}", status])

                    # Color code last row
                    last_row = ws.max_row
                    ws.cell(last_row, 5).fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Save file
            output_path = self.output_dir / filename
            wb.save(str(output_path))

            logger.info(f"Doğrulama raporu oluşturuldu: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Doğrulama raporu hatası: {str(e)}")
            raise

    def create_template_excel(
        self,
        template_fields: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        Create empty Excel template from field definitions

        Args:
            template_fields: Template field definitions
            filename: Optional custom filename

        Returns:
            Path to generated template file
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Şablon"

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"template_{timestamp}.xlsx"

            # Add headers
            headers = [field['field_name'] for field in template_fields]
            ws.append(headers)

            # Style header
            self._style_header_row(ws, len(headers))

            # Add example row
            example_row = []
            for field in template_fields:
                data_type = field.get('data_type', 'text')
                if data_type == 'date':
                    example_row.append("01/01/2024")
                elif data_type == 'number':
                    example_row.append("123,45")
                else:
                    example_row.append("Örnek metin")

            ws.append(example_row)

            # Auto-adjust columns
            self._auto_adjust_columns(ws)

            # Save file
            output_path = self.output_dir / filename
            wb.save(str(output_path))

            logger.info(f"Şablon Excel dosyası oluşturuldu: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Şablon oluşturma hatası: {str(e)}")
            raise
